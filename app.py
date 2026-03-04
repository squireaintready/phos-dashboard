"""
First Phosphate Corp (CSE: PHOS) — Financial Analysis Dashboard
Interactive Streamlit app with Excel view (default) + Dashboard toggle.
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl

st.set_page_config(
    page_title="PHOS Financial Dashboard",
    page_icon="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>⛏</text></svg>",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# --- SVG Icons ---
ICONS = {
    "overview": '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/></svg>',
    "financial": '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="12" y1="20" x2="12" y2="10"/><line x1="18" y1="20" x2="18" y2="4"/><line x1="6" y1="20" x2="6" y2="16"/></svg>',
    "fire": '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M8.5 14.5A2.5 2.5 0 0 0 11 12c0-1.38-.5-2-1-3-1.072-2.143-.224-4.054 2-6 .5 2.5 2 4.9 4 6.5 2 1.6 3 3.5 3 5.5a7 7 0 1 1-14 0c0-1.153.433-2.294 1-3a2.5 2.5 0 0 0 2.5 2.5z"/></svg>',
    "building": '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="4" y="2" width="16" height="20" rx="2" ry="2"/><path d="M9 22v-4h6v4"/><path d="M8 6h.01"/><path d="M16 6h.01"/><path d="M12 6h.01"/><path d="M12 10h.01"/><path d="M12 14h.01"/><path d="M16 10h.01"/><path d="M16 14h.01"/><path d="M8 10h.01"/><path d="M8 14h.01"/></svg>',
    "dollar": '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="12" y1="1" x2="12" y2="23"/><path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/></svg>',
    "shield": '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/></svg>',
    "users": '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/></svg>',
    "info": '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><line x1="12" y1="16" x2="12" y2="12"/><line x1="12" y1="8" x2="12.01" y2="8"/></svg>',
    "terminal": '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="4 17 10 11 4 5"/><line x1="12" y1="19" x2="20" y2="19"/></svg>',
}

# --- CSS ---
st.markdown("""
<style>
    /* Global */
    .block-container { padding-top: 1rem; max-width: 1400px; }
    
    /* Header */
    .main-header {
        display: flex; align-items: center; justify-content: space-between;
        padding: 12px 0; border-bottom: 1px solid #333; margin-bottom: 8px;
    }
    .main-header h1 { margin: 0; font-size: 22px; font-weight: 700; }
    .main-header .subtitle { color: #888; font-size: 13px; margin-top: 2px; }
    
    /* Tab bar at bottom */
    .tab-bar {
        position: fixed; bottom: 0; left: 0; right: 0; z-index: 9999;
        background: #0e1117; border-top: 1px solid #333;
        display: flex; justify-content: center; gap: 0;
        padding: 0; overflow-x: auto;
    }
    .tab-bar a {
        display: flex; flex-direction: column; align-items: center; gap: 3px;
        padding: 8px 14px; color: #888; text-decoration: none;
        font-size: 11px; white-space: nowrap; transition: color 0.15s;
        border-top: 2px solid transparent;
    }
    .tab-bar a:hover { color: #e94560; }
    .tab-bar a.active { color: #e94560; border-top-color: #e94560; }
    .tab-bar a svg { stroke: currentColor; }
    
    /* Excel-style table */
    .excel-table {
        width: 100%; border-collapse: collapse; font-size: 13px;
        font-family: 'Segoe UI', system-ui, sans-serif;
    }
    .excel-table th {
        background: #1a1a2e; color: #ccc; font-weight: 600;
        padding: 8px 12px; text-align: left; border: 1px solid #333;
        position: sticky; top: 0; z-index: 10;
    }
    .excel-table td {
        padding: 6px 12px; border: 1px solid #262640;
        color: #ddd; vertical-align: top;
    }
    .excel-table tr:nth-child(even) td { background: rgba(26, 26, 46, 0.3); }
    .excel-table tr:hover td { background: rgba(233, 69, 96, 0.08); }
    .excel-table .section-row td {
        background: #16213e !important; font-weight: 700;
        color: #e94560; border-bottom: 2px solid #e94560;
        padding: 10px 12px;
    }
    .excel-table .header-row td {
        background: #1a1a2e !important; font-weight: 600;
        color: #aaa; font-size: 12px;
    }
    .excel-table .number { text-align: right; font-variant-numeric: tabular-nums; }
    .excel-table .negative { color: #e94560; }
    .excel-table .positive { color: #53d769; }
    
    /* Metric cards */
    .metric-row { display: flex; gap: 12px; margin: 16px 0; flex-wrap: wrap; }
    .metric-card {
        flex: 1; min-width: 140px; background: #1a1a2e;
        border: 1px solid #0f3460; border-radius: 8px; padding: 16px;
    }
    .metric-card .label { font-size: 12px; color: #888; margin-bottom: 4px; }
    .metric-card .value { font-size: 22px; font-weight: 700; color: #e94560; }
    .metric-card .delta { font-size: 12px; color: #888; margin-top: 2px; }
    
    /* Star rating */
    .stars { color: #e94560; letter-spacing: 2px; }
    
    /* View toggle */
    .view-toggle {
        display: inline-flex; background: #1a1a2e; border-radius: 6px;
        border: 1px solid #333; overflow: hidden;
    }
    .view-toggle button {
        padding: 6px 16px; border: none; background: transparent;
        color: #888; cursor: pointer; font-size: 13px; transition: all 0.15s;
    }
    .view-toggle button.active {
        background: #e94560; color: white;
    }
    
    /* Bottom padding for tab bar */
    .main .block-container { padding-bottom: 80px; }
    
    /* Fix streamlit radio buttons to look like tabs */
    div[data-testid="stHorizontalBlock"] { gap: 0 !important; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# DATA LOADING
# ============================================================
@st.cache_data
def load_all_sheets():
    """Load all sheets as DataFrames preserving structure."""
    wb = openpyxl.load_workbook("data.xlsx", data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        sheets[name] = data
    return sheets


def sheet_to_html_table(data, max_cols=None):
    """Convert sheet data to styled HTML table."""
    if not data:
        return "<p>No data</p>"
    
    # Find actual max columns
    actual_max = max(len(row) for row in data)
    cols = max_cols or actual_max
    
    html = '<div style="overflow-x:auto; max-height: 70vh; overflow-y: auto;"><table class="excel-table">'
    
    for i, row in enumerate(data):
        # Pad row
        padded = list(row) + [None] * (cols - len(row))
        padded = padded[:cols]
        
        # Detect row type
        non_none = [v for v in padded if v is not None]
        is_section = (len(non_none) == 1 and i > 0 and 
                     isinstance(non_none[0], str) and 
                     (non_none[0].startswith("SECTION") or non_none[0].isupper()) and
                     len(non_none[0]) > 10)
        
        is_header = (i < 3) or (len(non_none) >= 2 and all(isinstance(v, str) for v in non_none) and
                    i > 0 and any(kw in str(non_none[0]).lower() for kw in ['name', 'line item', 'category', 'ratio', 'metric', 'scenario', 'company']))
        
        # First row = title
        if i == 0:
            title = non_none[0] if non_none else ""
            html += f'<tr class="section-row"><td colspan="{cols}">{title}</td></tr>'
            continue
        
        # Second/third row = subtitle
        if i in (1, 2):
            sub = non_none[0] if non_none else ""
            if sub:
                html += f'<tr class="header-row"><td colspan="{cols}" style="font-style:italic">{sub}</td></tr>'
            continue
        
        # Empty row
        if not non_none:
            continue
        
        # Section header
        if is_section:
            html += f'<tr class="section-row"><td colspan="{cols}">{non_none[0]}</td></tr>'
            continue
        
        # Detect if this looks like a column header row
        if is_header and i > 2:
            html += '<tr>'
            for v in padded:
                html += f'<th>{v if v is not None else ""}</th>'
            html += '</tr>'
            continue
        
        # Regular data row
        html += '<tr>'
        for j, v in enumerate(padded):
            if v is None:
                html += '<td></td>'
            elif isinstance(v, (int, float)):
                cls = "number"
                if v < 0:
                    cls += " negative"
                elif v > 0 and j > 0:
                    cls += " positive"
                # Format numbers
                if abs(v) >= 1_000_000:
                    formatted = f"${v/1_000_000:,.1f}M"
                elif abs(v) >= 1_000:
                    formatted = f"${v/1_000:,.1f}K"
                elif abs(v) < 1 and v != 0:
                    formatted = f"{v:.4f}"
                else:
                    formatted = f"{v:,.0f}"
                html += f'<td class="{cls}">{formatted}</td>'
            else:
                s = str(v)
                # Detect star ratings
                if '★' in s:
                    html += f'<td><span class="stars">{s}</span></td>'
                else:
                    html += f'<td>{s}</td>'
        html += '</tr>'
    
    html += '</table></div>'
    return html


# ============================================================
# DASHBOARD RENDERERS
# ============================================================

def dashboard_overview(data):
    st.markdown("""
    <div class="metric-row">
        <div class="metric-card"><div class="label">Share Price</div><div class="value">C$1.05</div><div class="delta">+193% 1yr</div></div>
        <div class="metric-card"><div class="label">Market Cap</div><div class="value">~C$158M</div><div class="delta">FD ~C$202M</div></div>
        <div class="metric-card"><div class="label">Cash (Q3)</div><div class="value">C$20.0M</div><div class="delta">Post-raise</div></div>
        <div class="metric-card"><div class="label">PEA NPV (8%)</div><div class="value">C$1.59B</div><div class="delta">After-tax</div></div>
        <div class="metric-card"><div class="label">Mkt Cap / NPV</div><div class="value">~11%</div><div class="delta">LFP premium</div></div>
    </div>
    """, unsafe_allow_html=True)
    
    left, right = st.columns(2)
    with left:
        quarters = ["Q4 FY24", "Q1 FY25", "Q2 FY25", "Q3 FY25", "Q4 FY25", "Q1 FY26", "Q2 FY26", "Q3 FY26"]
        cash = [7496238, 1651673, 410444, 149983, 1873550, 3173855, 7590632, 19983238]
        fig = go.Figure(go.Bar(x=quarters, y=[c/1e6 for c in cash],
            marker_color=['#e94560' if c < 1e6 else '#0f3460' for c in cash],
            text=[f"${c/1e6:.1f}M" for c in cash], textposition='outside'))
        fig.update_layout(title="Cash Position Over Time", yaxis_title="C$ Millions",
            template="plotly_dark", height=350, margin=dict(t=40, b=40), yaxis=dict(gridcolor='#222'))
        st.plotly_chart(fig, use_container_width=True)
    
    with right:
        shares = [73786772, 74867570, 76103368, 77198802, 89947551, 97023899, 123546512, 151220000]
        fig = go.Figure(go.Scatter(x=quarters, y=[s/1e6 for s in shares],
            mode='lines+markers+text', line=dict(color='#e94560', width=3), marker=dict(size=8),
            text=[f"{s/1e6:.0f}M" for s in shares], textposition='top center'))
        fig.update_layout(title="Share Dilution", yaxis_title="Shares Outstanding (M)",
            template="plotly_dark", height=350, margin=dict(t=40, b=40), yaxis=dict(gridcolor='#222'))
        st.plotly_chart(fig, use_container_width=True)
    
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("#### Bull Case — C$1.84/sh")
        st.success("FS confirms PEA. OEM partnership secured. LFP demand accelerates. 20% NPV discount.")
    with c2:
        st.markdown("#### Base Case — C$1.01/sh")
        st.info("Current market pricing (~11% NPV). Drill program on schedule. LFP premium maintained.")
    with c3:
        st.markdown("#### Bear Case — C$0.28/sh")
        st.error("Resource downgrade in PFS. Capex funding failure. 3% NPV discount (typical PEA explorer).")


def dashboard_financial(data):
    periods = ["FY2024", "FY2025", "Q1 FY26", "Q2 FY26", "Q3 FY26"]
    assets = [12995758, 7452772, 8735500, 14682667, 25126133]
    liabs = [3684258, 1063172, 809820, 758567, 1235763]
    equity = [9311500, 6389600, 7925680, 13924100, 23890370]
    
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Bar(name="Total Assets", x=periods, y=[a/1e6 for a in assets], marker_color='#0f3460'))
    fig.add_trace(go.Bar(name="Liabilities", x=periods, y=[l/1e6 for l in liabs], marker_color='#e94560'))
    fig.add_trace(go.Scatter(name="Equity", x=periods, y=[e/1e6 for e in equity],
        mode='lines+markers', line=dict(color='#53d769', width=3)), secondary_y=True)
    fig.update_layout(title="Balance Sheet Trend", template="plotly_dark", height=380,
        barmode='group', margin=dict(t=40, b=40), legend=dict(orientation="h", y=1.12))
    fig.update_yaxes(title_text="C$ Millions", secondary_y=False, gridcolor='#222')
    fig.update_yaxes(title_text="Equity (C$M)", secondary_y=True, gridcolor='#222')
    st.plotly_chart(fig, use_container_width=True)
    
    loss_q = ["Q4 FY24", "Q1 FY25", "Q2 FY25", "Q3 FY25", "Q4 FY25", "Q1 FY26", "Q2 FY26"]
    loss_v = [3764747, 4185217, 170997, 1681986, 1589214, 1748338, 2006009]
    fig = go.Figure(go.Bar(x=loss_q, y=[v/1e6 for v in loss_v],
        marker_color=['#e94560' if v > 2e6 else '#ff8a5c' for v in loss_v],
        text=[f"${v/1e6:.1f}M" for v in loss_v], textposition='outside'))
    fig.update_layout(title="Net Loss by Quarter", yaxis_title="C$ Millions",
        template="plotly_dark", height=320, margin=dict(t=40, b=40), yaxis=dict(gridcolor='#222'))
    st.plotly_chart(fig, use_container_width=True)


def dashboard_cashburn(data):
    st.markdown("""
    <div class="metric-row">
        <div class="metric-card"><div class="label">Quarterly Burn</div><div class="value">~C$2.7M</div></div>
        <div class="metric-card"><div class="label">Est. Cash (Jan 2026)</div><div class="value">~C$32.6M</div></div>
        <div class="metric-card"><div class="label">Runway (Base)</div><div class="value">~30 months</div></div>
    </div>
    """, unsafe_allow_html=True)
    
    quarters = ["Q4 FY24", "Q1 FY25", "Q2 FY25", "Q3 FY25", "Q4 FY25", "Q1 FY26", "Q2 FY26"]
    burn = [3764747, 4185217, 170997, 1681986, 1589214, 1748338, 2006009]
    cum_dil = [1.5, 3.1, 4.6, 21.9, 31.5, 67.4, 104.9]
    
    left, right = st.columns(2)
    with left:
        fig = go.Figure(go.Bar(x=quarters, y=[b/1e6 for b in burn], marker_color='#e94560',
            text=[f"${b/1e6:.1f}M" for b in burn], textposition='outside'))
        fig.update_layout(title="Quarterly Net Loss", template="plotly_dark", height=320,
            margin=dict(t=40, b=40), yaxis=dict(gridcolor='#222'))
        st.plotly_chart(fig, use_container_width=True)
    
    with right:
        fig = go.Figure(go.Scatter(x=quarters, y=cum_dil, mode='lines+markers+text',
            fill='tozeroy', line=dict(color='#e94560', width=3),
            text=[f"{d:.0f}%" for d in cum_dil], textposition='top center',
            fillcolor='rgba(233,69,96,0.15)'))
        fig.update_layout(title="Cumulative Dilution from FY24", template="plotly_dark", height=320,
            margin=dict(t=40, b=40), yaxis=dict(gridcolor='#222'))
        st.plotly_chart(fig, use_container_width=True)
    
    st.warning("Capex Gap: PEA estimates C$675M+ construction cost vs ~C$33M current cash.")


def dashboard_peers(data):
    companies = ["PHOS", "DAN\n(Arianne)", "NMG\n(Graphite)", "PMET\n(Lithium)"]
    npv_pct = [11, 1.5, 34, 37]
    colors = ['#e94560', '#0f3460', '#53d769', '#ff8a5c']
    fig = go.Figure(go.Bar(x=companies, y=npv_pct, marker_color=colors,
        text=[f"{p}%" for p in npv_pct], textposition='outside', textfont=dict(size=16)))
    fig.update_layout(title="Market Cap as % of NPV — LFP Supply Chain",
        template="plotly_dark", height=400, margin=dict(t=40, b=40), yaxis=dict(gridcolor='#222'))
    st.plotly_chart(fig, use_container_width=True)
    st.info("PHOS is the cheapest entry point in North American LFP supply chain at ~11% of NPV vs 34% (NMG) and 37% (PMET).")


def dashboard_valuation(data):
    npv = 1_590_000_000
    npv_pct = st.slider("NAV Discount %", 3, 30, 11, 1)
    implied_nav = npv * (npv_pct / 100)
    implied_price_fd = implied_nav / 192_400_000
    upside = (implied_price_fd / 1.05) - 1
    
    c1, c2, c3 = st.columns(3)
    c1.metric("Implied NAV", f"C${implied_nav/1e6:.0f}M")
    c2.metric("Implied Price (FD)", f"C${implied_price_fd:.2f}",
              f"{'↑' if upside > 0 else '↓'} {abs(upside)*100:.0f}% vs C$1.05")
    assess = "Very Cheap" if npv_pct <= 3 else "Cheap" if npv_pct <= 5 else "Fair" if npv_pct <= 7 else "~Current" if npv_pct <= 12 else "Rich"
    c3.metric("Assessment", assess)
    
    scenarios = ["Bear (3%)", "Base (11%)", "Bull (20%)"]
    prices = [0.28, 1.01, 1.84]
    fig = go.Figure(go.Bar(x=scenarios, y=prices, marker_color=['#e94560','#0f3460','#53d769'],
        text=[f"C${p:.2f}" for p in prices], textposition='outside', textfont=dict(size=18)))
    fig.add_hline(y=1.05, line_dash="dash", line_color="white", annotation_text="Current: C$1.05")
    fig.update_layout(title="Scenario Analysis", template="plotly_dark", height=350,
        margin=dict(t=40, b=40), yaxis=dict(gridcolor='#222'))
    st.plotly_chart(fig, use_container_width=True)


def dashboard_risk(data):
    st.markdown("#### Key Findings")
    findings = pd.DataFrame({
        "Area": ["Financial Statements", "Cash Burn", "Cash Burn", "Peer Analysis", "Valuation"],
        "Finding": ["Rapid BS growth from equity raises", "Cash runway adequate if burn stable",
                    "Massive capex gap unfunded", "Only pure-play LFP phosphate junior",
                    "Significant upside in bull scenario"],
        "Metric": ["Assets: C$7.5M → C$25.1M", "~24 months at C$2.7M/qtr",
                   "C$675M+ vs C$33M cash", "No direct comparable", "Bear C$0.28 → Bull C$1.84"],
        "Risk": ["Medium", "Medium", "HIGH", "Low", "Medium"],
    })
    st.dataframe(findings, use_container_width=True, hide_index=True)
    
    st.markdown("#### Upcoming Catalysts")
    catalysts = [
        "PFS/FS Completion (2026-2027) — Major re-rating event",
        "OEM Partnership / Offtake agreement",
        "Federal Critical Minerals Grant — C$4.9M from NRCan",
        "30,000m Drill Program — Resource upgrade from 83% Inferred",
        "ADR Listing — US market access expansion",
    ]
    for c in catalysts:
        st.markdown(f"- {c}")


def dashboard_mgmt(data):
    st.markdown("""
    <div class="metric-row">
        <div class="metric-card"><div class="label">CEO Open Market Buying</div><div class="value">C$1.8M</div><div class="delta">2,872,000 shares since May 2023</div></div>
        <div class="metric-card"><div class="label">CEO Cash Salary</div><div class="value">$0</div><div class="delta">100% equity comp</div></div>
        <div class="metric-card"><div class="label">Board Cash Fees</div><div class="value">$0</div><div class="delta">100% RSU comp since Sep 2023</div></div>
        <div class="metric-card"><div class="label">Overall Rating</div><div class="value" style="color:#53d769">Above Avg</div><div class="delta">for stage</div></div>
    </div>
    """, unsafe_allow_html=True)
    
    cats = ["Executive\nLeadership", "Board\nIndependence", "Technical\nExpertise", 
            "Strategic\nRelationships", "Compensation\nAlignment", "Insider\nConviction", "Operational\nReadiness"]
    ratings = [4, 3, 5, 5, 5, 5, 3]
    colors = ['#53d769' if r >= 4 else '#ff8a5c' if r >= 3 else '#e94560' for r in ratings]
    
    fig = go.Figure(go.Bar(x=cats, y=ratings, marker_color=colors,
        text=[f"{'★'*r}{'☆'*(5-r)}" for r in ratings], textposition='outside'))
    fig.update_layout(title="Management Assessment by Category", template="plotly_dark",
        height=380, margin=dict(t=40, b=40), yaxis=dict(gridcolor='#222', range=[0, 5.5], dtick=1))
    st.plotly_chart(fig, use_container_width=True)


DASHBOARD_RENDERERS = {
    "Company Overview": dashboard_overview,
    "Financial Statements": dashboard_financial,
    "Cash Burn Analysis": dashboard_cashburn,
    "Peer Analysis": dashboard_peers,
    "Valuation Model": dashboard_valuation,
    "Analysis & Summary": dashboard_risk,
    "Management & Governance": dashboard_mgmt,
}


# ============================================================
# MAIN APP
# ============================================================

sheets = load_all_sheets()
sheet_names = list(sheets.keys())

# Tab config with icons
TAB_CONFIG = [
    ("Company Overview", "info"),
    ("Financial Statements", "financial"),
    ("Cash Burn Analysis", "fire"),
    ("Peer Analysis", "building"),
    ("Valuation Model", "dollar"),
    ("Analysis & Summary", "shield"),
    ("Management & Governance", "users"),
    ("Claude Log", "terminal"),
]

# --- Header ---
st.markdown("""
<div class="main-header">
    <div>
        <h1>First Phosphate Corp.</h1>
        <div class="subtitle">CSE: PHOS | OTCQX: FRSPF | Pre-Revenue LFP Battery Phosphate Developer</div>
    </div>
</div>
""", unsafe_allow_html=True)

# --- Navigation tabs ---
tab_labels = [name for name, _ in TAB_CONFIG]
selected_tab = st.radio("", tab_labels, horizontal=True, label_visibility="collapsed")

# --- View toggle ---
col_spacer, col_toggle = st.columns([4, 1])
with col_toggle:
    view_mode = st.toggle("Dashboard View", value=False)

st.markdown("---")

# --- Content ---
if selected_tab in sheets:
    data = sheets[selected_tab]
elif selected_tab == "Company Overview" and "Company Overview" in sheets:
    data = sheets["Company Overview"]
else:
    data = []

if view_mode and selected_tab in DASHBOARD_RENDERERS:
    # Dashboard view
    DASHBOARD_RENDERERS[selected_tab](data)
else:
    # Excel view (default)
    if data:
        html = sheet_to_html_table(data)
        st.markdown(html, unsafe_allow_html=True)
    else:
        st.info("No data available for this sheet.")

# --- Footer ---
st.markdown("---")
st.caption("Built by Samuel Jo | Data: SEDAR+ filings | Not investment advice | [GitHub](https://github.com/squireaintready/phos-dashboard)")
